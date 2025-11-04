import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import time
from datetime import datetime
import logging
import threading
from typing import Dict, Optional
import os
import random

# Selenium-specific imports
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

# --- SELENIUM-BASED SCRAPER LOGIC (FINAL VERSION) ---

class BITResultsScraper:
    """The core scraper class, with complete data parsing for the results page."""
    def __init__(self, gui_logger=None):
        self.driver = None
        self.gui_logger = gui_logger
        self.base_url = "https://ioncudos.in/bit_online_results/"

    def log(self, message):
        if self.gui_logger:
            self.gui_logger(message)
        else:
            print(f"[DEBUG] {message}")

    def initialize_driver(self):
        try:
            self.log("Initializing Chrome browser...")
            options = webdriver.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
            
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.log("✅ Chrome browser initialized successfully.")
            return True
        except Exception as e:
            self.log(f"❌ Critical Error: Could not initialize Chrome driver: {e}")
            return False

    def quit_driver(self):
        if self.driver:
            self.log("Closing Chrome browser.")
            self.driver.quit()
            self.driver = None
            
    def format_date(self, date_input) -> str:
        if pd.isna(date_input): raise ValueError("Date of birth is missing")
        if isinstance(date_input, str):
            formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']
            for fmt in formats:
                try:
                    return datetime.strptime(date_input, fmt).strftime('%d-%m-%Y')
                except ValueError: continue
            raise ValueError(f"Unable to parse date string: {date_input}")
        elif isinstance(date_input, datetime): return date_input.strftime('%d-%m-%Y')
        else: raise ValueError(f"Unsupported date format: {type(date_input)}")

    def get_result_data(self, usn: str, dob: str) -> Optional[Dict]:
        try:
            formatted_dob = self.format_date(dob)
            
            self.log(f"Navigating to {self.base_url} to ensure a clean state...")
            self.driver.get(self.base_url)

            usn_input = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "usn")))
            dob_input = self.driver.find_element(By.ID, "dob")
            
            usn_input.clear()
            usn_input.send_keys(usn)

            self.log(f"Setting DOB to {formatted_dob} using JavaScript...")
            self.driver.execute_script("arguments[0].value = arguments[1];", dob_input, formatted_dob)
            
            self.log("Triggering 'change' event on DOB field...")
            self.driver.execute_script("var e = new Event('change', {bubbles: true}); arguments[0].dispatchEvent(e);", dob_input)
            time.sleep(0.5)

            self.log("Forcibly enabling the submit button using JavaScript...")
            submit_button = self.driver.find_element(By.ID, "result_submit")
            self.driver.execute_script("arguments[0].removeAttribute('disabled');", submit_button)
            
            self.log("Button is now enabled. Clicking submit.")
            submit_button.click()
            
            self.log("Waiting for results to load...")
            WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.ID, "stud_name"))
            )
            
            self.log("Results loaded. Parsing page content...")
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            return self.parse_html_result(soup, usn)

        except TimeoutException as e:
            self.log(f"Page timed out. Likely invalid USN/DOB or page didn't load. Error: {e.msg}")
            return None
        except Exception as e:
            self.log(f"An unexpected error occurred: {e}")
            return None

    def parse_html_result(self, soup: BeautifulSoup, usn: str) -> Optional[Dict]:
        """Parses all student details, including the full 10-column subject table."""
        try:
            result_data = {'USN': usn, 'Name': '', 'Program': '', 'Semester': '', 'SGPA': '', 'CGPA': '', 'Subjects': []}
            
            # Scrape general student info
            result_data['Name'] = getattr(soup.find('span', {'id': 'stud_name'}), 'text', '').strip()
            if not result_data['Name']: return None # Fail early if no name is found
            
            result_data['Program'] = getattr(soup.find('span', {'id': 'stud_pgm'}), 'text', '').strip()
            result_data['Semester'] = getattr(soup.find('span', {'id': 'stud_sem'}), 'text', '').strip()
            result_data['SGPA'] = getattr(soup.find('span', {'id': 'stud_sgpa'}), 'text', '').strip()
            result_data['CGPA'] = getattr(soup.find('span', {'id': 'stud_cgpa'}), 'text', '').strip()
            
            # Scrape the subjects table
            result_table_div = soup.find('div', {'id': 'result_table'})
            if result_table_div and result_table_div.find('table'):
                rows = result_table_div.find('table').find_all('tr')[1:]
                for row in rows:
                    cols = row.find_all('td')
                    
                    # <-- UPDATED: Check for 10 columns to include 'Credits Earned' -->
                    if len(cols) >= 10:
                        subject_data = {
                            'Course_Name': cols[0].get_text(strip=True),
                            'Course_Code': cols[1].get_text(strip=True),
                            'Credits': cols[2].get_text(strip=True),
                            'CIE': cols[3].get_text(strip=True),
                            'SEE': cols[4].get_text(strip=True),
                            'Total_Marks': cols[5].get_text(strip=True),
                            'Letter_Grade': cols[6].get_text(strip=True),
                            'Credits_Earned': cols[7].get_text(strip=True), # <-- NEWLY ADDED
                            'Grade_Points': cols[8].get_text(strip=True),   # <-- Index updated
                            'Credit_Points': cols[9].get_text(strip=True)  # <-- Index updated
                        }
                        result_data['Subjects'].append(subject_data)

            self.log(f"Successfully parsed: {result_data['Name']} (SGPA: {result_data['SGPA']})")
            return result_data
        except Exception as e:
            self.log(f"💥 Error during HTML parsing: {e}")
            return None


# --- TKINTER GUI APPLICATION (Unchanged) ---
class BITResultsScraperGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("BIT Results Scraper (Selenium Version)")
        self.root.geometry("850x750")
        self.input_file_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()
        self.delay_var = tk.DoubleVar(value=2.0)
        self.progress_var = tk.DoubleVar()
        self.is_scraping = False
        self.scraper = BITResultsScraper(gui_logger=self.log_message)
        self.setup_gui()
    def setup_gui(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        title_label = ttk.Label(main_frame, text="BIT Results Scraper (Selenium)", font=('Arial', 18, 'bold'))
        title_label.pack(pady=(0, 20))
        io_frame = ttk.LabelFrame(main_frame, text="File Settings", padding="10")
        io_frame.pack(fill=tk.X, pady=10)
        io_frame.columnconfigure(1, weight=1)
        ttk.Label(io_frame, text="Input Excel File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(io_frame, textvariable=self.input_file_path, width=60).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(io_frame, text="Browse...", command=self.browse_input_file).grid(row=0, column=2, padx=5)
        ttk.Label(io_frame, text="Output Folder:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(io_frame, textvariable=self.output_folder_path, width=60).grid(row=1, column=1, sticky="ew", padx=5)
        ttk.Button(io_frame, text="Browse...", command=self.browse_output_folder).grid(row=1, column=2, padx=5)
        options_frame = ttk.LabelFrame(main_frame, text="Options", padding="10")
        options_frame.pack(fill=tk.X, pady=10)
        ttk.Label(options_frame, text="Delay (seconds):").pack(side=tk.LEFT, padx=(5,0))
        delay_scale = ttk.Scale(options_frame, from_=1.0, to=10.0, variable=self.delay_var, orient=tk.HORIZONTAL, length=200)
        delay_scale.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        delay_label = ttk.Label(options_frame, width=5)
        self.delay_var.trace_add("write", lambda *args: delay_label.config(text=f"{self.delay_var.get():.1f}s"))
        delay_label.config(text=f"{self.delay_var.get():.1f}s")
        delay_label.pack(side=tk.LEFT, padx=(0, 10))
        paned_window = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        paned_window.pack(fill=tk.BOTH, expand=True, pady=10)
        preview_frame = ttk.LabelFrame(paned_window, text="File Preview", padding="5")
        self.preview_text = scrolledtext.ScrolledText(preview_frame, height=8, width=70, wrap=tk.WORD)
        self.preview_text.pack(fill=tk.BOTH, expand=True)
        paned_window.add(preview_frame, weight=1)
        log_frame = ttk.LabelFrame(paned_window, text="Log", padding="5")
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=70, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        paned_window.add(log_frame, weight=3)
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=10)
        progress_frame.columnconfigure(0, weight=1)
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.progress_label = ttk.Label(progress_frame, text="Ready", width=25)
        self.progress_label.grid(row=0, column=1, sticky=tk.E)
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        self.start_button = ttk.Button(button_frame, text="Start Scraping", command=self.start_scraping)
        self.start_button.pack(side=tk.LEFT, padx=10)
        self.stop_button = ttk.Button(button_frame, text="Stop", command=self.stop_scraping, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Clear Log", command=self.clear_log).pack(side=tk.LEFT, padx=10)
    def start_scraping(self):
        if not self.input_file_path.get() or not self.output_folder_path.get():
            messagebox.showerror("Error", "Please select both an input file and an output folder.")
            return
        self.is_scraping = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.progress_var.set(0)
        self.scraping_thread = threading.Thread(target=self.scraping_worker, daemon=True)
        self.scraping_thread.start()
    def scraping_worker(self):
        if not self.scraper.initialize_driver():
            self.root.after(0, lambda: messagebox.showerror("Driver Error", "Failed to start Chrome. Check logs for details."))
            self.root.after(0, self.reset_ui_state)
            return
        try:
            df = pd.read_excel(self.input_file_path.get())[['USN', 'DOB', 'StudentName', 'Gender']]
            total_records = len(df)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(self.output_folder_path.get(), f"BIT_Results_Selenium_{timestamp}.xlsx")
            self.log_message(f"Starting to scrape {total_records} records...")
            all_results, failed_usns = [], []
            for index, row in df.iterrows():
                if not self.is_scraping:
                    self.log_message("Scraping stopped by user.")
                    break
                usn, dob, name, gender = str(row['USN']).strip(), row['DOB'], row['StudentName'], row['Gender']
                self.progress_var.set(((index + 1) / total_records) * 100)
                self.progress_label.config(text=f"{index + 1}/{total_records}: {usn}")
                result = self.scraper.get_result_data(usn, dob)
                if result and result.get('Name'):
                    result['Gender'] = gender  # Add gender from input file
                    all_results.append(result)
                else: failed_usns.append(usn)
                time.sleep(self.delay_var.get())
            if all_results: 
                self.save_results_to_excel(all_results, output_file)
                self.log_message(f"\n📁 OUTPUT FILE SAVED AT:")
                self.log_message(f"📄 {output_file}")
            else:
                self.log_message("❌ No results to save. All scraping attempts failed.")
            
            final_msg = f"Scraping complete. Successful: {len(all_results)}, Failed: {len(failed_usns)}."
            if all_results:
                final_msg += f"\n\n📁 File saved at:\n{output_file}"
            self.log_message("=" * 40 + f"\n{final_msg}\n" + "=" * 40)
            if self.is_scraping: self.root.after(0, lambda: messagebox.showinfo("Complete", final_msg))
        except Exception as e:
            self.log_message(f"A critical error occurred: {e}")
            self.root.after(0, lambda: messagebox.showerror("Critical Error", str(e)))
        finally:
            self.scraper.quit_driver()
            self.root.after(0, self.reset_ui_state)
    def browse_input_file(self):
        file_path = filedialog.askopenfilename(title="Select Input Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.input_file_path.set(file_path)
            self.preview_file(file_path)
            if not self.output_folder_path.get(): self.output_folder_path.set(os.path.dirname(file_path))
    def browse_output_folder(self):
        folder_path = filedialog.askdirectory(title="Select Output Folder")
        if folder_path: self.output_folder_path.set(folder_path)
    def preview_file(self, file_path):
        self.preview_text.delete(1.0, tk.END)
        try:
            df = pd.read_excel(file_path)
            required_cols = ['USN', 'DOB', 'StudentName', 'Gender']
            missing_cols = [col for col in required_cols if col not in df.columns]
            info = f"File: {os.path.basename(file_path)}\nRows: {len(df)}\n\n"
            if missing_cols: info += f"⚠️ ERROR: Missing required columns: {', '.join(missing_cols)}"
            else: info += "✅ All required columns (USN, DOB, StudentName, Gender) found.\n\n--- First 5 Rows ---\n" + df[required_cols].head(5).to_string(index=False)
            self.preview_text.insert(tk.END, info)
        except Exception as e: self.preview_text.insert(tk.END, f"Error reading file: {e}")
    def stop_scraping(self):
        if self.is_scraping:
            self.is_scraping = False
            self.log_message("Stopping... please wait for the current request to finish.")
            self.stop_button.config(state=tk.DISABLED)
    def save_results_to_excel(self, results, output_file):
        self.log_message(f"Saving {len(results)} results to {os.path.basename(output_file)}...")
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Get all unique subject codes from first result (assuming all students have same subjects)
            subject_codes = []
            if results and results[0].get('Subjects'):
                subject_codes = [subj.get('Course_Code', '') for subj in results[0]['Subjects']]
            
            # Build header rows
            # Row 1: Main headers (Subject codes span across CIE, SEE, Total, Grade)
            header_row1 = ['USN', 'Name', 'Gender']
            for code in subject_codes:
                header_row1.extend([code, '', '', ''])  # Subject code spans 4 columns
            header_row1.extend(['SGPA', 'CGPA', 'Final Grade'])
            
            # Row 2: Sub-headers (CIE, SEE, Total, Grade under each subject)
            header_row2 = ['', '', '']  # Empty under USN, Name, Gender
            for _ in subject_codes:
                header_row2.extend(['CIE', 'SEE', 'Total', 'GR'])
            header_row2.extend(['', '', ''])  # Empty under SGPA, CGPA, Final Grade
            
            # Write headers
            ws.append(header_row1)
            ws.append(header_row2)
            
            # Merge cells for main headers
            from openpyxl.styles import Alignment, Font
            col_offset = 4  # Starting column for first subject (after USN, Name, Gender)
            for i, code in enumerate(subject_codes):
                start_col = col_offset + (i * 4)
                end_col = start_col + 3
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            
            # Make first row bold and centered
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for cell in ws[2]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data rows
            for res in results:
                row_data = [
                    res.get('USN', ''),
                    res.get('Name', ''),
                    res.get('Gender', '')
                ]
                
                # Add subject data in order
                has_failure = False
                subjects_dict = {subj.get('Course_Code', ''): subj for subj in res.get('Subjects', [])}
                
                for code in subject_codes:
                    subject = subjects_dict.get(code, {})
                    row_data.extend([
                        subject.get('CIE', ''),
                        subject.get('SEE', ''),
                        subject.get('Total_Marks', ''),
                        subject.get('Letter_Grade', '')
                    ])
                    
                    # Check for failure
                    if subject.get('Letter_Grade', '').upper() == 'F':
                        has_failure = True
                
                # Add SGPA, CGPA, Final Grade
                row_data.extend([
                    res.get('SGPA', ''),
                    res.get('CGPA', ''),
                    'Fail' if has_failure else 'Pass'
                ])
                
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_file)
            
            self.log_message(f"✅ Successfully saved file with {len(results)} records and multi-level headers.")
            self.log_message(f"✅ File location: {output_file}")
            
            # Check if file exists
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file) / 1024  # Size in KB
                self.log_message(f"✅ File size: {file_size:.2f} KB")
        except Exception as e: 
            self.log_message(f"❌ Error saving Excel file: {e}")
            import traceback
            self.log_message(traceback.format_exc())
    def reset_ui_state(self):
        self.is_scraping = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.progress_label.config(text="Complete")
    def clear_log(self): self.log_text.delete(1.0, tk.END)
    def log_message(self, message: str):
        log_entry = f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n"
        self.root.after(0, self._append_to_log, log_entry)
    def _append_to_log(self, text: str):
        self.log_text.insert(tk.END, text)
        self.log_text.see(tk.END)


if __name__ == "__main__":
    import argparse
    import json
    import sys
    
    parser = argparse.ArgumentParser(description='Autonomous Results Scraper')
    parser.add_argument('--url', type=str, help='Results URL', required=False)
    parser.add_argument('--workers', type=int, default=1, help='Number of parallel workers (not used yet)')
    parser.add_argument('--students', type=str, help='JSON string with student list [{usn, dob}, ...]', required=False)
    
    args = parser.parse_args()
    
    # If CLI args provided, use CLI mode
    if args.url and args.students:
        print("="*70)
        print("🏫 AUTONOMOUS RESULTS SCRAPER (CLI Mode)")
        print("="*70)
        print()
        
        try:
            students = json.loads(args.students)
            print(f"📋 Loading {len(students)} students from arguments...")
            
            # Initialize scraper
            def cli_logger(msg):
                print(msg)
            
            scraper = BITResultsScraper(gui_logger=cli_logger)
            
            if not scraper.initialize_driver():
                print("❌ Failed to initialize Chrome driver")
                sys.exit(1)
            
            all_results = []
            failed_usns = []
            success_count = 0
            failed_count = 0
            
            print()
            print("="*70)
            print("STARTING SCRAPE...")
            print("="*70)
            print()
            
            for idx, student in enumerate(students, 1):
                usn = student.get('usn', '').strip()
                dob = student.get('dob', '').strip()
                
                if not usn or not dob:
                    print(f"❌ SKIP: {usn} - Missing USN or DOB")
                    failed_usns.append(usn)
                    failed_count += 1
                    continue
                
                print(f"[{idx}/{len(students)}] Processing: {usn}")
                
                result = scraper.get_result_data(usn, dob)
                
                if result and result.get('Name'):
                    all_results.append(result)
                    success_count += 1
                    print(f"✅ SUCCESS: {usn} - {result.get('Name')}")
                else:
                    failed_usns.append(usn)
                    failed_count += 1
                    print(f"❌ FAIL: {usn} - Could not fetch results")
                
                time.sleep(1)  # Small delay between requests
            
            scraper.quit_driver()
            
            print()
            print("="*70)
            print("🎉 SCRAPING COMPLETE!")
            print("="*70)
            print(f"✅ Success: {success_count}")
            print(f"❌ Failed: {failed_count}")
            print("="*70)
            
        except json.JSONDecodeError as e:
            print(f"❌ Invalid JSON in --students argument: {e}")
            sys.exit(1)
        except Exception as e:
            print(f"❌ Critical error: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    else:
        # GUI mode (original behavior)
        root = tk.Tk()
        app = BITResultsScraperGUI(root)
        root.mainloop()
